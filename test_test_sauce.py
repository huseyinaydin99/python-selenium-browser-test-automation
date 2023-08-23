from unittest import TestCase
from selenium import webdriver
from time import sleep
from unittest import TestCase
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
import pytest
from pathlib import Path
from datetime import date
import openpyxl
#import constants.globalConstants as const
from constants import globalConstants as const
class TestTest_Sauce(TestCase):
    def setUp(self):
        self.driver = webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get(const.URL)
        # Bugünün tarhini al. Bu tarih adında bir klasör var mı onu bir kontrol et. Yoksa o tarih isminde yeni bir klasör oluştur.
        self.folderPath = str(date.today())
        Path(self.folderPath).mkdir(exist_ok=True)


    def tearDown(self):
        self.driver.quit()

    def getData():
        excelFile = openpyxl.load_workbook("data/invalid_login.xlsx")
        selectedSheet = excelFile["Sheet1"]
        totalRows = selectedSheet.max_rows
        data = []
        for i in range(2, totalRows + 1):
            username = selectedSheet.cell(i, 1).value
            password = selectedSheet.cell(i, 2).value
            tupleData = (username, password)
            data.append(tupleData)
        return data

    @pytest.mark.parameterize("username", "password", getData())
    def test_invalid_login(self, username, password):
        self.driver.get(const.URL)
        self.waitForElementVisible((By.ID, "user-name"))
        usernameInput = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID, "password"))
        passwordInput = self.driver.find_element(By.ID, "password")
        usernameInput.send_keys("1")
        passwordInput.send_keys("1")
        loginButton = self.driver.find_element(By.XPATH, "//*[@id='login-button']")
        loginButton.click()
        errorMessage = self.driver.find_element(By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test-invalid-login-{username}-{password}.png")
        print(errorMessage)
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"
        sleep(2)
    def waitForElementVisible(self, locator, timeout=5):
        WebDriverWait(self.driver, timeout).until(expected_conditions.visibility_of_element_located(locator))


    def test_valid_login(self):
        WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.ID, "user-name")))
        usernameInput = self.driver.find_element(By.ID, "user-name")
        WebDriverWait(self.driver, 5).until(expected_conditions.visibility_of_element_located((By.ID, "password")))
        passwordInput = self.driver.find_element(By.ID, "password")
        self.driver.execute_script("window.scrollTo(0, 500)")
        # action chains
        actions = ActionChains(self.driver)
        actions.send_keys_to_element(usernameInput, "standard_user")
        actions.send_keys_to_element(passwordInput, "secret_sauce")
        actions.perform()
        loginBtn = self.driver.find_element(By.ID, "login-button")
        loginBtn.click()
